"""
Excel Helper - Tạo template Excel trống và export kết quả scan
"""

import os
from pathlib import Path


class ExcelHelper:
    """Helper để tạo và xử lý file Excel"""
    
    @staticmethod
    def create_template_excel(output_path="batch_videos_template.xlsx"):
        """Tạo file Excel template trống để user paste link"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Videos"
            
            # Set column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            
            # Header row
            headers = ["Watch full rescues", "Title (Optional)", "Status"]
            
            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Add 20 empty rows for user to fill
            for row in range(2, 22):
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 3:  # Status column
                        cell.alignment = Alignment(horizontal="center")
            
            # Add instruction sheet
            ws_info = wb.create_sheet("Instructions")
            ws_info.column_dimensions['A'].width = 80
            
            instructions = [
                "📋 HƯỚNG DẪN SỬ DỤNG",
                "",
                "1. Cột 'Watch full rescues':",
                "   - Paste link video vào đây (mỗi dòng 1 link)",
                "   - Ví dụ: https://vimeo.com/123456789",
                "",
                "2. Cột 'Title (Optional)':",
                "   - Tên bài viết (nếu không điền sẽ tự động lấy từ video)",
                "",
                "3. Cột 'Status':",
                "   - Sẽ được cập nhật tự động sau khi scan",
                "",
                "4. Sau khi điền xong:",
                "   - Lưu file",
                "   - Import vào tool bằng nút 'Import Excel'",
                "   - Tool sẽ tự động scan tất cả link",
                "",
                "5. Kết quả:",
                "   - File Excel mới sẽ được tạo với đầy đủ thông tin",
                "   - Bao gồm: Title, Thumbnail, Embed Code, Video ID",
            ]
            
            for row_num, instruction in enumerate(instructions, 1):
                cell = ws_info.cell(row=row_num, column=1)
                cell.value = instruction
                if instruction.startswith("📋"):
                    cell.font = Font(bold=True, size=14)
                elif instruction.startswith(("1.", "2.", "3.", "4.", "5.")):
                    cell.font = Font(bold=True, size=11)
            
            # Save
            wb.save(output_path)
            print(f"[EXCEL] ✅ Template created: {output_path}")
            return output_path
            
        except ImportError:
            print("[EXCEL] ⚠️ openpyxl not installed. Install: pip install openpyxl")
            return None
        except Exception as e:
            print(f"[EXCEL] ❌ Error creating template: {e}")
            return None
    
    @staticmethod
    def read_excel(file_path):
        """Read video links from Excel"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            videos = []
            
            # Find URL column
            url_col = None
            title_col = None
            
            for col_num, cell in enumerate(ws[1], 1):
                if cell.value:
                    cell_lower = str(cell.value).lower()
                    if 'watch' in cell_lower or 'url' in cell_lower or 'link' in cell_lower:
                        url_col = col_num
                    if 'title' in cell_lower or 'name' in cell_lower:
                        title_col = col_num
            
            if not url_col:
                return None, "Không tìm thấy cột URL/Link"
            
            print(f"[EXCEL] URL Column: {url_col}, Title Column: {title_col}")
            
            # Read data rows
            for row_num in range(2, ws.max_row + 1):
                url_cell = ws.cell(row=row_num, column=url_col)
                url = url_cell.value
                
                if url and isinstance(url, str):
                    url = url.strip()
                    if url:
                        title = ""
                        if title_col:
                            title_cell = ws.cell(row=row_num, column=title_col)
                            title = title_cell.value or ""
                        
                        videos.append({
                            'url': url,
                            'title': title or 'Auto-generated',
                            'row': row_num
                        })
            
            print(f"[EXCEL] ✅ Read {len(videos)} videos from Excel")
            return videos, None
            
        except ImportError:
            return None, "openpyxl not installed"
        except Exception as e:
            return None, f"Error reading Excel: {e}"
    
    @staticmethod
    def export_results(videos, output_path="batch_videos_results.xlsx"):
        """Export scan results to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Set column widths
            ws.column_dimensions['A'].width = 50  # URL
            ws.column_dimensions['B'].width = 30  # Title
            ws.column_dimensions['C'].width = 50  # Thumbnail
            ws.column_dimensions['D'].width = 80  # Embed Code
            ws.column_dimensions['E'].width = 15  # Video ID
            ws.column_dimensions['F'].width = 12  # Status
            
            # Header
            headers = ["URL", "Title", "Thumbnail", "Embed Code", "Video ID", "Status"]
            
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Data rows
            for row_num, video in enumerate(videos, 2):
                ws.cell(row=row_num, column=1).value = video.get('url', '')
                ws.cell(row=row_num, column=2).value = video.get('title', '')
                ws.cell(row=row_num, column=3).value = video.get('thumbnail', '')
                ws.cell(row=row_num, column=4).value = video.get('embed_code', '')
                ws.cell(row=row_num, column=5).value = video.get('video_id', '')
                ws.cell(row=row_num, column=6).value = video.get('status', 'pending')
                
                # Style data rows
                for col in range(1, 7):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    if col == 6:  # Status
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Save
            wb.save(output_path)
            print(f"[EXCEL] ✅ Results exported: {output_path}")
            return output_path
            
        except ImportError:
            print("[EXCEL] ⚠️ openpyxl not installed")
            return None
        except Exception as e:
            print(f"[EXCEL] ❌ Error exporting: {e}")
            return None


# Test
if __name__ == "__main__":
    helper = ExcelHelper()
    
    # Create template
    print("Creating template...")
    template_path = helper.create_template_excel()
    
    if template_path:
        print(f"✅ Template created at: {template_path}")
